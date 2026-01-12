import sys
import openpyxl
from tkinter import messagebox, filedialog
import time
import xml.etree.ElementTree as ET
import os
import re
import win32com.client
import openpyxl
import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog
import time

# Conexión a SAP / SAP connection

Excel_File = filedialog.askopenfilename(title="Select Excel File Template",filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
wb = openpyxl.load_workbook(Excel_File)
VariantROW = 2

VariantCOLCH = -1
VariantCOLVAL = 0
sheet = wb['XML_Output']
XML_File = filedialog.askdirectory(title="Select XML File")
start = time.time()
files = 0
for archivo in os.listdir(XML_File):
            if archivo.endswith(".xml"):
                files += 1
                VariantROW = 2
                VariantCOLCH += 2
                VariantCOLVAL += 2
                ruta_completa = os.path.join(XML_File, archivo)
                sheet.cell(1,VariantCOLCH).value = "File name:"
                sheet.cell(1,VariantCOLVAL).value = os.path.basename(ruta_completa)
                with open(ruta_completa,'r', encoding='UTF-8') as f:
                    raw_content = f.read()

                clean_content = raw_content.replace('\\','')

                clean_content = clean_content.strip()
                if clean_content.startswith('"') and clean_content.endswith('"'):
                    clean_content = clean_content[1:-1]
                clean_content = re.sub(r'^[^\<]*', '', clean_content)
                print(clean_content)
                XML_tree = ET.ElementTree(ET.fromstring(clean_content))
                #This script works for XML wit SAP XML Nodes and CDATA
                sap_xml_nodes = XML_tree.findall(".//SAP_XML")

                

                for node in sap_xml_nodes:
                    if node.text:
                        cdata_content = node.text.strip()

                        try:

                            sap_root = ET.fromstring(cdata_content)

                            cstics = sap_root.findall(".//CSTIC")

                            for cstic in cstics:

                                charc = cstic.attrib.get('CHARC')
                                value = cstic.attrib.get('VALUE')
                                print(f'CHARC: {charc} VALUE: {value}')
                                sheet.cell(VariantROW,VariantCOLCH).value = charc
                                sheet.cell(VariantROW,VariantCOLVAL).value = value
                                VariantROW +=1 
                                
                            wb.save(Excel_File)
                        except ET.ParseError as e:
                            messagebox.showinfo("Error",f"Error al procesar CDATA: {e}")
wb.save(Excel_File)

#Process for entering values to SAP
CisticROW = 2
CisticColumn = 1
ValueColumn = 2
while sheet.cell(CisticROW,CisticColumn) is not None:
     # Conexión a SAP / SAP connection
    sap_gui_auto = win32com.client.GetObject("SAPGUI")
    sap_gui_app = sap_gui_auto.GetScriptingEngine
    connection = sap_gui_app.Children(0)
    session = connection.Children(0)

    # Type KMAT Here
    Kmat = simpledialog.askstring("KMAT", "Enter KMAT:")
    # Type plant Here
    Plant = simpledialog.askstring("Plant", "Enter Plant:")
    # Enter the date for CU50
    Date = simpledialog.askstring("Date", "Enter ECN Date  MM/DD/YYYY:")

    session.findById("wnd[0]").maximize ()
    session.findById("wnd[0]/tbar[0]/okcd").text = "/ncs03"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/usr/ctxtRC29N-MATNR").text = Kmat
    session.findById("wnd[0]/usr/ctxtRC29N-WERKS").text = Plant
    session.findById("wnd[0]/usr/ctxtRC29N-STLAN").text = "3"
    session.findById("wnd[0]/usr/ctxtRC29N-STLAN").setFocus
    session.findById("wnd[0]/usr/ctxtRC29N-STLAN").caretPosition = 1
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/tbar[0]/okcd").text = "/ncu50"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/usr/ctxtRCUKO-DATUV").text = Date
    session.findById("wnd[0]/tbar[1]/btn[8]").press ()
    session.findById("wnd[1]/tbar[0]/btn[0]").press ()

    #Find Window, and select CHAR which was found
    def Select_FoundValue():
        session.findById("wnd[2]/usr/cntlGRID1/shellcont/shell").selectedRows = "0"
        session.findById("wnd[2]/usr/cntlGRID1/shellcont/shell").clickCurrentCell ()
        
    #Enter value in SAP from Excel
    def enter_value():
        session.findById("wnd[0]/usr/subCE_INSTANCE:SAPLCEI0:1105/tabsTABSTRIP_CHAR/tabpTAB1/ssubCHARACTERISTICS:SAPLCEI0:1400/tblSAPLCEI0CHARACTER_VALUES/ctxtRCTMS-MWERT[1,0]").text = str(sheet.cell(row=CisticROW,column=ValueColumn).value)
        session.findById("wnd[0]").sendVKey (0)

    #Move between tabs
    def tabs():
        session.findById(f"wnd[0]/usr/subCE_INSTANCE:SAPLCEI0:1105/tabsTABSTRIP_CHAR/tabpTAB{tab}").select()
        session.findById("wnd[0]/usr/subCE_INSTANCE:SAPLCEI0:1105/btnSUCHE").press ()
        session.findById("wnd[1]/usr/txtCLHP-CR_STATUS_TEXT").text = sheet.cell(row=CisticROW,column=CisticColumn).value 
        session.findById("wnd[1]/tbar[0]/btn[0]").press () 
            

    Num_Tab = simpledialog.askinteger("Tabs", "Enter the Qty of Tabs in the Interface")
    tab = 1

    while sheet.cell(CisticROW,CisticColumn).value is not None:
         
        if tab == 0:
            tab = 1
        if Num_Tab == 1 and sheet.cell(row=CisticROW,column=ValueColumn).value is not None:
            session.findById("wnd[0]/usr/subCE_INSTANCE:SAPLCEI0:1105/btnSUCHE").press ()
            session.findById("wnd[1]/usr/txtCLHP-CR_STATUS_TEXT").text = sheet.cell(row=CisticROW,column=CisticColumn).value
            session.findById("wnd[1]/tbar[0]/btn[0]").press ()
            
            try:    
                    Result1 = ""
                    Result2 = ""  
                    Result1 = session.findById("wnd[1]/usr/txtMESSTXT1").text
                    Result2 = session.findById("wnd[1]/usr/txtMESSTXT2").text
                    Result = Result1 + " " + Result2
                    if "Not Displayed" in Result:
                        CisticROW += 1

            except:
                pass

            try:
                Select_FoundValue()

            except:
                session.findById("wnd[0]/usr/subCE_INSTANCE:SAPLCEI0:0105/subCHARACTERISTICS:SAPLCEI0:1400/tblSAPLCEI0CHARACTER_VALUES/ctxtRCTMS-MWERT[1,0]").text = str(sheet.cell(row=CisticROW,column=ValueColumn).value)
                session.findById("wnd[0]").sendVKey (0)
                CisticROW += 1
                continue

        else:

                while tab <= Num_Tab:

                    if sheet.cell(row=CisticROW,column=ValueColumn).value is not None:
                        if tab < Num_Tab:
                            tabs()
                            try:
                                Result1 = ""
                                Result2 = "" 
                                Result1 = session.findById("wnd[1]/usr/txtMESSTXT1").text
                                Result2 = session.findById("wnd[1]/usr/txtMESSTXT2").text
                                Result = Result1 + " " + Result2
                                session.findById("wnd[1]").close ()
                                if "Not Displayed" in Result:
                                    CisticROW += 1
                                    break
                            except: 
                                pass
                            tab +=1
                        
                        elif tab == Num_Tab:
                            tabs()
                            try:
                                Result1 = ""
                                Result2 = "" 
                                Result1 = session.findById("wnd[1]/usr/txtMESSTXT1").text
                                Result2 = session.findById("wnd[1]/usr/txtMESSTXT2").text
                                Result = Result1 + " " + Result2
                                if "Not Displayed" in Result:
                                    CisticROW += 1
                            except: 
                                pass
                            tab =1  

                        try:
                            Select_FoundValue()
                            session.findById("wnd[0]").sendVKey (0)
                        
                        except:
                            pass

                        try:
                            #Close window where it was not found CHAR
                            session.findById("wnd[1]/tbar[0]/btn[0]").press ()

                        except:
                            enter_value()
                            if tab <= Num_Tab:
                                tab -=1
                            elif tab == 1:
                                tab = Num_Tab
                            CisticROW += 1


