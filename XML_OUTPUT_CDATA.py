import sys
import openpyxl
from tkinter import messagebox, filedialog
import time
import xml.etree.ElementTree as ET
import os
import re
import win32com.client
import openpyxl
import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog
import time

# Conexión a SAP / SAP connection

Excel_File = filedialog.askopenfilename(title="Select Excel File Template",filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
wb = openpyxl.load_workbook(Excel_File)
VariantROW = 2

VariantCOLCH = -1
VariantCOLVAL = 0
sheet = wb['XML_Output']
XML_File = filedialog.askdirectory(title="Select XML File")
start = time.time()
files = 0
for archivo in os.listdir(XML_File):
            if archivo.endswith(".xml"):
                files += 1
                VariantROW = 2
                VariantCOLCH += 2
                VariantCOLVAL += 2
                ruta_completa = os.path.join(XML_File, archivo)
                sheet.cell(1,VariantCOLCH).value = "File name:"
                sheet.cell(1,VariantCOLVAL).value = os.path.basename(ruta_completa)
                with open(ruta_completa,'r', encoding='UTF-8') as f:
                    raw_content = f.read()

                clean_content = raw_content.replace('\\','')

                clean_content = clean_content.strip()
                if clean_content.startswith('"') and clean_content.endswith('"'):
                    clean_content = clean_content[1:-1]
                clean_content = re.sub(r'^[^\<]*', '', clean_content)
                print(clean_content)
                XML_tree = ET.ElementTree(ET.fromstring(clean_content))
                #This script works for XML wit SAP XML Nodes and CDATA
                sap_xml_nodes = XML_tree.findall(".//SAP_XML")

                

                for node in sap_xml_nodes:
                    if node.text:
                        cdata_content = node.text.strip()

                        try:

                            sap_root = ET.fromstring(cdata_content)

                            cstics = sap_root.findall(".//CSTIC")

                            for cstic in cstics:

                                charc = cstic.attrib.get('CHARC')
                                value = cstic.attrib.get('VALUE')
                                print(f'CHARC: {charc} VALUE: {value}')
                                sheet.cell(VariantROW,VariantCOLCH).value = charc
                                sheet.cell(VariantROW,VariantCOLVAL).value = value
                                VariantROW +=1 
                                
                            wb.save(Excel_File)
                        except ET.ParseError as e:
                            messagebox.showinfo("Error",f"Error al procesar CDATA: {e}")
wb.save(Excel_File)

